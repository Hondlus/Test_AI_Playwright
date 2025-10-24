from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
import os, sys, re, json
import pandas as pd
from neepshop_UI import CustomDialog
from datetime import datetime, timedelta
import logging
from markitdown import MarkItDown
import zipfile
import requests
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


def setup_logging():
    """配置日志系统"""
    # 创建logs目录
    log_dir = 'logs'
    if not os.path.exists(log_dir):
        os.makedirs(log_dir)

    # 设置日志文件名（按日期）
    log_filename = datetime.now().strftime('neepshop_%Y%m%d-%H%M%S.log')
    log_filepath = os.path.join(log_dir, log_filename)

    # 配置日志格式
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_filepath, encoding='utf-8'),
            logging.StreamHandler(sys.stdout)  # 同时输出到控制台
        ]
    )

    return logging.getLogger(__name__)

cookie_json = 'neepshop.json'
success_url = 'https://www.neep.shop/'
login_url = 'https://cooperation.ceic.com/login/index?client_id=oauth-neep&redirect_uri=https%3A%2F%2Fwww.neep.shop%2Frest%2Fstsso&response_type=code'
excel_url = 'https://www.neep.shop/html/portal/notice.html?type=enquiryOrderAnnc&nodeurl=callback_list_enquiry_order&noticeMoreUrl=https://gd-prod.cn-beijing.oss.aliyuncs.com/upload/cms/column/inquireListOne/index.html&pageTag=undefined&menu_code=&parent_menu_code=&root_menu_code='
pdf_url = 'https://www.neep.shop/dist/index.html#/purchaserNoticeIndex#/purchaserNoticeIndex?autoId=290201'
fabu_time_file = 'public_time.txt'
create_time_file = 'create_time.txt'
logger = setup_logging()

api_key = "fastgpt-fCYCYBicNtBob8rzrytnbB60rivhEduElK0wWzBVCE2AB3RxJqKc0kZ9sURcPaNc"
base_url = "http://192.168.50.81:3100/ragai"
workflow_id = "68e74de6fd26a9e519813e3c"
chat_id = "chat_id"
file_id = True
pattern = r'\{[^{}]*\}'
separator = "-" * 60


def write_excel(xmmc, bjrzgtj, xjfs, wzfl, fwsj, bjjzsj, fbsj, kw, is_down_pdf):
    try:
        data = {
            '项目名称': [xmmc],
            '报价人资格条件': [bjrzgtj],
            '询价方式': [xjfs],
            '物资分类': [wzfl],
            '服务时间': [fwsj],
            '报价截止时间': [bjjzsj],
            '发布时间': [fbsj],
            '是否下载pdf': [is_down_pdf]
        }
        file_path = os.path.join(os.getcwd(), kw) + '/' + '01_招标网站智能搜索结果.xlsx'
        try:
            existing_df = pd.read_excel(file_path)
            new_df = pd.DataFrame(data)
            # 将新数据追加到现有的DataFrame中
            combined_df = pd.concat([existing_df, new_df], ignore_index=True)
            # 将合并后的数据写回Excel文件
            combined_df.to_excel(file_path, index=False, sheet_name='信息表')
            logger.info("数据已通过Pandas成功追加并保存！")
        except FileNotFoundError:
            logger.info("文件不存在，创建新文件")
            df = pd.DataFrame(data)
            df.to_excel(file_path, index=False, sheet_name='信息表')
            logger.info("Excel文件已生成！")
    except Exception as e:
        logger.error(f"写入Excel文件时发生错误: {str(e)}")
        raise


def write_excel2(xmmc, ai_read_text, kw):
    try:
        data = {'文件名': [xmmc]}
        # 定义新的输出顺序
        new_order = ['业务承接判定', '承接判定说明', '项目名称', '项目类型', '项目建设内容']
        # 获取原始字典的所有键
        original_keys = list(ai_read_text.keys())
        # 从原始键中移除已经放在开头的键
        remaining_keys = [key for key in original_keys if key not in new_order]
        # 构建新的键顺序
        final_order = new_order + remaining_keys
        # 按照新顺序创建有序字典
        data.update({key: [ai_read_text[key]] for key in final_order})

        file_path = os.path.join(os.getcwd(), kw) + '/' + '02_招标文件智能解析结果.xlsx'
        try:
            existing_df = pd.read_excel(file_path)
            new_df = pd.DataFrame(data)
            # 将新数据追加到现有的DataFrame中
            combined_df = pd.concat([existing_df, new_df], ignore_index=True)
            # 将合并后的数据写回Excel文件
            combined_df.to_excel(file_path, index=False, sheet_name='信息表')
            logger.info("数据已通过Pandas成功追加并保存！")
        except FileNotFoundError:
            logger.info("excel文件不存在，创建新文件")
            df = pd.DataFrame(data)
            df.to_excel(file_path, index=False, sheet_name='信息表')
            logger.info("Excel文件已生成！")
    except Exception as e:
        logger.error(f"写入Excel文件时发生错误: {str(e)}")
        raise

    # existing_df = pd.read_excel(file_path)
    # trans_df = existing_df.T.reset_index()
    # trans_df.to_excel(file_path, index=False, sheet_name='信息表')

def format_excel_file(file_path):
    # 加载工作簿
    wb = load_workbook(file_path)
    ws = wb.active

    # --------------设置第一行列宽、背景颜色、文字颜色-----------------
    # 冻结第一列和第二列（A列和B列）
    # 冻结窗格在C1单元格，即保持A列和B列可见
    ws.freeze_panes = 'C1'

    column_widths = [19, 17, 37, 23, 20, 76, 21, 16, 21, 23,
                     15, 22, 15, 19, 15, 21, 21, 24, 15, 16,
                     18, 26, 24, 24, 24, 22, 23, 19, 19, 28,
                     23, 13, 30, 25, 25, 24, 24, 36]

    for i, width in enumerate(column_widths, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # 设置亮蓝色背景和白色字体
    light_blue_fill = PatternFill(start_color="00B0F0",
                                  end_color="00B0F0",
                                  fill_type="solid")

    white_font = Font(color="FFFFFF", bold=True, size=14)  # 白色字体

    for cell in ws[1]:  # 第一行所有单元格
        cell.fill = light_blue_fill
        cell.font = white_font

    # --------------设置所有行垂直居中、文字居左、自动换行-----------------
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            # 设置单元格对齐方式
            cell.alignment = Alignment(
                vertical='center',  # 垂直居中
                horizontal='left',  # 水平居左
                wrap_text=True  # 自动换行
            )
    # 保存文件
    wb.save(file_path)
    logger.info("Excel文件格式设置完成！")


def update_content(project_name_list, iframe_locator, context, keyword, old_time):
    project_name = None
    baojiarenzigetiaojian = None
    xunjiafangshi = None
    wuzifenlei = None
    fuwushijian = None
    baojiajiezhishijian = None
    tmp_new_time = None

    logger.info(f"开始处理关键词 '{keyword}' 的项目列表，共 {len(project_name_list)} 个项目")

    for i in range(len(project_name_list)):
        try:
            logger.info(f"处理第 {i + 1} 个项目: {project_name_list[i]}")

            link_locator = iframe_locator.get_by_role('link', name=project_name_list[i])
            href = link_locator.get_attribute('href')
            if not href:
                logger.warning(f"项目 '{project_name_list[i]}' 没有找到链接，跳过")
                continue
            page3 = context.new_page()
            page3.goto(href)
            items = page3.query_selector_all('div.notice-detail div.item')
            items2 = page3.query_selector_all('div.notice-db div.item-db div.item')
            public_times = page3.query_selector_all('div.content_right div.right-main-content div.help-detail-title div.title-other')

            for public_time in public_times:
                if i == 0:
                    tmp_new_time = public_time.inner_text().strip()
                fabushijian = public_time.inner_text().strip()
                try:
                    new_time = datetime.strptime(fabushijian, "%Y-%m-%d %H:%M:%S")
                    logger.info(f"项目发布时间: {fabushijian}")
                except ValueError as e:
                    logger.error(f"时间格式解析错误: {fabushijian}, 错误: {e}")
                    page3.close()
                    continue

            if old_time < new_time:
                logger.info("发现新数据，开始提取信息")
                for item in items:
                    text_content = item.inner_text()
                    if "项目名称" in text_content:
                        project_name = text_content.split("：", 1)[1].strip()
                        logger.info(f"项目名称: {project_name}")
                    elif "报价人资格条件" in text_content:
                        baojiarenzigetiaojian = text_content.split("：", 1)[1].strip()
                        logger.info(f"报价人资格条件: {baojiarenzigetiaojian}")
                    elif "公开询价" in text_content:
                        xunjiafangshi = text_content.split("：", 1)[1].strip()
                        logger.info(f"询价方式: {xunjiafangshi}")
                    elif "服务类->综合服务" in text_content:
                        wuzifenlei = text_content.split("：", 1)[1].strip()
                        logger.info(f"物资分类: {wuzifenlei}")

                for item2 in items2:
                    text_content2 = item2.inner_text()
                    if "服务时间" in text_content2 or "交货时间" in text_content2:
                        fuwushijian = text_content2.split("：", 1)[1].strip()
                        logger.info(f"交货时间或服务时间: {fuwushijian}")
                    elif "报价截止时间" in text_content2:
                        baojiajiezhishijian = text_content2.split("：", 1)[1].strip()
                        logger.info(f"报价截止时间: {baojiajiezhishijian}")

                if project_name and baojiarenzigetiaojian and xunjiafangshi and wuzifenlei and fuwushijian and baojiajiezhishijian:
                    logger.info("项目信息完整，符合条件，开始下载PDF")
                    # ------------------------下载pdf-----------------------
                    page3.goto(pdf_url)
                    try:
                        page3.get_by_role("textbox", name="请输入采购单名称").click(timeout=6000)
                        page3.get_by_role("textbox", name="请输入采购单名称").fill(project_name)
                        page3.get_by_role("button", name="搜索").click()
                        logger.info(f"已搜索采购单: {project_name}")
                    except PlaywrightTimeoutError:
                        logger.warning("账户未登录，无法下载PDF文件")
                        dialog = CustomDialog("账户登录提示", "账户未登录,不能下载pdf文件,请登陆账户,再重新执行程序")
                        dialog.exec()

                    try:
                        # 设置导航等待超时
                        with page3.expect_navigation(timeout=10000):
                            page3.get_by_role("row", name="序号 采购单名称 采购单编号 收到的澄清 日期/周期 发布时间 报价(名)截止时间 采购机构 采购类别").get_by_label("").check()
                            page3.get_by_role("button", name="我要参与").click()
                            page3.get_by_role("button", name="确定").click()

                        logger.info("页面发生了跳转, 加载页面A:报编")
                        try:
                            page3.get_by_role("button", name="关闭").wait_for(state="visible", timeout=10000)
                            page3.get_by_role("button", name="关闭").click()
                            logger.info("已关闭弹窗")
                        except PlaywrightTimeoutError:
                            logger.info("页面没有关闭按钮，直接下载")

                        download_button = page3.get_by_role("button", name=" 下载采购文件")
                        with page3.expect_download() as download_info:
                            download_button.click()

                        # 获取下载对象
                        download = download_info.value
                        # 等待下载文件完成并获取建议的文件名
                        suggested_filename = download.suggested_filename
                        file_path = os.path.join(os.path.join(os.getcwd(), keyword, '附件', '01_AI未解析'), suggested_filename)
                        # 将文件保存到指定路径（如果已有同名文件，可能会覆盖）
                        download.save_as(file_path)
                        logger.info(f"PDF文件已下载到: {file_path}")
                        write_excel(project_name, baojiarenzigetiaojian, xunjiafangshi, wuzifenlei, fuwushijian,
                                    baojiajiezhishijian, fabushijian, keyword, '是')

                    except Exception as e:
                        try:
                            with context.expect_page(timeout=10000) as new_page_info:
                                page3.get_by_role("row", name="序号 采购单名称 采购单编号 收到的澄清 日期/周期 发布时间 报价(名)截止时间 采购机构 采购类别").get_by_label("").check()
                                page3.get_by_role("button", name="我要参与").click()
                                page3.get_by_role("button", name="确定").click()
                            new_page = new_page_info.value
                            logger.info(f"没有发生页面跳转，可能是新标签页: {e}")
                            logger.info("加载页面B:供应商询比价管理")

                            try:
                                new_page.wait_for_selector('a.fileOperation-btn', timeout=10000)
                                download_elements = new_page.query_selector_all('a.fileOperation-btn')
                                logger.info(f"找到 {len(download_elements)} 个下载链接")

                                for download_element in download_elements:
                                    with new_page.expect_download() as download_info:
                                        download_element.click()
                                        # 获取下载对象
                                        download = download_info.value
                                        # 等待下载文件完成并获取建议的文件名
                                        suggested_filename = download.suggested_filename
                                        file_path = os.path.join(
                                            os.path.join(os.getcwd(), keyword, '附件', '01_AI未解析', project_name),
                                            suggested_filename)
                                        # 将文件保存到指定路径（如果已有同名文件，可能会覆盖）
                                        download.save_as(file_path)
                                        logger.info(f"WORD文件已下载到: {file_path}")
                                write_excel(project_name, baojiarenzigetiaojian, xunjiafangshi, wuzifenlei,
                                            fuwushijian, baojiajiezhishijian, fabushijian, keyword, '是')
                            except Exception as e:
                                logger.info(f"页面下载按钮版面变化导致错误: {e}")
                            new_page.close()

                        except Exception as e:
                            logger.info(f"未知错误: {e}")
                            logger.info(f"未搜索到 {project_name} 相关下载文件网页")
                            write_excel(project_name, baojiarenzigetiaojian, xunjiafangshi, wuzifenlei, fuwushijian,
                                        baojiajiezhishijian, fabushijian, keyword, '否')
                            continue
                    # ------------------------下载pdf----------------------
                    logger.info("新数据更新成功")
                else:
                    logger.warning("项目信息不完整，不符合条件")
            else:
                logger.info("数据无需更新，跳过处理")
                page3.close()
                break
            page3.close()

        except Exception as e:
            logger.error(f"处理项目 '{project_name_list[i]}' 时发生错误: {str(e)}")
            if 'page3' in locals():
                page3.close()
            continue

    logger.info(f"关键词 '{keyword}' 处理完成")
    return tmp_new_time


def download_excel_pdf(main_folder, fabu_time_file, cookie_json, excel_url):
    logger.info("开始执行下载任务")
    with sync_playwright() as p:
        try:
            bro = p.chromium.launch(headless=False, slow_mo=1000)
            context = bro.new_context(storage_state=cookie_json)
            page = context.new_page()
            logger.info("浏览器启动成功")

            for keyword in main_folder:
                logger.info(f"开始处理关键词: {keyword}")
                main_path = os.path.join(os.getcwd(), keyword)
                if not os.path.exists(main_path):
                    os.makedirs(os.path.join(main_path, '附件', '01_AI未解析'))
                    os.makedirs(os.path.join(main_path, '附件', '02_AI已解析(不可承接)'))
                    os.makedirs(os.path.join(main_path, '附件', '03_AI已解析(可承接)'))
                    logger.info(f"创建主文件夹: {main_path}")
                else:
                    logger.info(f"主文件夹已存在: {main_path}")

                page.goto(excel_url)
                logger.info(f"已打开excel项目地址页面: {excel_url}")
                page.locator("#notice iframe").content_frame.get_by_role("textbox", name="项目名称").click()
                page.locator("#notice iframe").content_frame.get_by_role("textbox", name="项目名称").fill(keyword)
                page.locator("#notice iframe").content_frame.get_by_text("搜索").click()
                logger.info(f"已搜索关键词: {keyword}")

                iframe_locator = page.frame_locator("iframe[src='https://gd-prod.cn-beijing.oss.aliyuncs.com/upload/cms/column/inquireListOne/index.html']")
                project_name_list = iframe_locator.locator('[class="c_href"]').all_text_contents()
                # next_page = iframe_locator.locator('//*[@id="next_page"]').inner_text()
                pagenum = iframe_locator.locator('//*[@id="pageNum"]').inner_text()
                total_page = int(pagenum.lstrip('共').rstrip('页'))
                logger.info(f"找到 {len(project_name_list)} 个项目，共 {total_page} 页")

                if len(project_name_list) == 0:
                    logger.info("没有网页数据")
                    continue
                else:
                    if not os.path.exists(os.path.join(os.getcwd(), keyword, fabu_time_file)):
                        logger.info("时间文件不存在，创建初始时间")
                        # 获取当前日期（时间部分设为00:00:00）
                        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
                        # 计算5天前的日期
                        five_days_ago = today - timedelta(days=5)
                        # 按照指定格式输出
                        today_str = today.strftime("%Y-%m-%d %H:%M:%S")
                        five_days_ago_str = five_days_ago.strftime("%Y-%m-%d %H:%M:%S")
                        logger.info(f"设置初始时间: {five_days_ago_str}")

                        # 创建标记文件
                        with open(os.path.join(os.getcwd(), keyword, fabu_time_file), 'w') as f:
                            f.write(five_days_ago_str)

                        with open(os.path.join(os.getcwd(), keyword, fabu_time_file), 'r') as file:
                            lines = file.readlines()
                            old_fabu_time = lines[0].strip() if lines else None  # 处理空文件
                            old_time = datetime.strptime(old_fabu_time, "%Y-%m-%d %H:%M:%S")
                            logger.info(f"读取上次更新时间: {old_fabu_time}")

                        # 正常遍历内容
                        tmp_time = update_content(project_name_list, iframe_locator, context, keyword, old_time)
                        # 下一页功能实现
                        for i in range(total_page - 1):
                            logger.info(f"处理第 {i + 2} 页")
                            page.locator("#notice iframe").content_frame.get_by_text("下一页").click()
                            iframe_locator = page.frame_locator("iframe[src='https://gd-prod.cn-beijing.oss.aliyuncs.com/upload/cms/column/inquireListOne/index.html']")
                            project_name_list = iframe_locator.locator('[class="c_href"]').all_text_contents()
                            update_content(project_name_list, iframe_locator, context, keyword, old_time)
                        with open(os.path.join(os.getcwd(), keyword, fabu_time_file), 'w', encoding='utf-8') as file:
                            file.write(tmp_time)
                        logger.info(f"更新时间文件: {tmp_time}")

                    # 如果不为空，则取出txt中旧时间，进行对比
                    else:
                        with open(os.path.join(os.getcwd(), keyword, fabu_time_file), 'r') as file:
                            lines = file.readlines()
                            old_fabu_time = lines[0].strip() if lines else None  # 处理空文件
                            old_time = datetime.strptime(old_fabu_time, "%Y-%m-%d %H:%M:%S")

                        # 正常遍历内容
                        tmp_time2 = update_content(project_name_list, iframe_locator, context, keyword, old_time)
                        # 下一页功能实现
                        for i in range(total_page - 1):
                            page.locator("#notice iframe").content_frame.get_by_text("下一页").click()
                            iframe_locator = page.frame_locator(
                                "iframe[src='https://gd-prod.cn-beijing.oss.aliyuncs.com/upload/cms/column/inquireListOne/index.html']")
                            project_name_list = iframe_locator.locator('[class="c_href"]').all_text_contents()
                            update_content(project_name_list, iframe_locator, context, keyword, old_time)
                        with open(os.path.join(os.getcwd(), keyword, fabu_time_file), 'w', encoding='utf-8') as file:
                            file.write(tmp_time2)
            logger.info("所有关键词处理完成")

        except Exception as e:
            logger.error(f"执行下载任务时发生错误: {str(e)}")
            raise
        finally:
            page.close()
            context.close()
            bro.close()
            logger.info("浏览器已关闭")


def main(keywords_list):
    is_login_require = False
    main_folder = keywords_list
    # 账户cookie登录逻辑
    if os.path.exists(os.path.join(os.getcwd(), cookie_json)):
        with sync_playwright() as p:
            # 启动浏览器，假设get_browser_object函数已实现或直接启动
            browser = p.chromium.launch(headless=False)  # 设为True则无头模式运行
            context = browser.new_context(storage_state=cookie_json)
            page = context.new_page()
            page.goto(login_url)
            # page.wait_for_load_state("networkidle")
            try:
                page.wait_for_url(success_url)
                dialog = CustomDialog("账户登录提示", "账号登陆成功，关闭弹窗，继续执行程序")
                dialog.exec()
                page.close()
                context.close()
                browser.close()
                is_login_require = True
                logger.info("账户登录成功，开始执行程序")
            except Exception as e:
                logger.info("cookie过期失效,请重新登录账户")
                page.close()
                context.close()
                browser.close()
                dialog = CustomDialog("账户登录提示", "cookie过期失效,请重新登录账户")
                dialog.exec()
                os.remove(cookie_json)
                is_login_require = False
            finally:
                page.close()
                context.close()
                browser.close()

    else:
        dialog = CustomDialog("账户登录提示", "请登陆账户")
        dialog.exec()
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=False)
            context = browser.new_context()
            page = context.new_page()
            page.goto(login_url)
            try:
                page.wait_for_url(success_url, timeout=5*60*1000)
                dialog = CustomDialog("账户登录提示", "账户登录成功,关闭弹窗,继续执行程序")
                dialog.exec()
                storage = context.storage_state(path=cookie_json)
                page.close()
                context.close()
                browser.close()
                is_login_require = True
                logger.info("账户cookie创建成功,开始执行程序")
            except Exception as e:
                logger.info("登录超时或出错")
                page.close()
                context.close()
                browser.close()
                is_login_require = False
                dialog = CustomDialog("账户登录提示", "登陆账户超时或出错,重新登陆")
                dialog.exec()
            finally:
                page.close()
                context.close()
                browser.close()

    if is_login_require:
        dialog = CustomDialog("账户登陆状态", "账户登陆成功True,开始收集数据和下载pdf")
        dialog.exec()
        download_excel_pdf(main_folder, fabu_time_file, cookie_json, excel_url)
        dialog = CustomDialog("账户登陆状态", "收集数据和下载pdf完成")
        dialog.exec()
    else:
        dialog = CustomDialog("账户登陆状态", "账户登陆失败False")
        dialog.exec()


def main2(keywords_list):
    dialog = CustomDialog("程序执行中,请等待...", "程序执行中，请等待...")
    dialog.show()

    for keyword in keywords_list:
        if not os.path.exists(os.path.join(keyword, '附件', '01_AI未解析')):
            logger.info(f"{keyword}-附件-01_AI未解析 文件夹不存在,请检查相关数据")
        else:
            if len(os.listdir(os.path.join(keyword, '附件', '01_AI未解析'))) == 0:
                logger.info(f"{keyword}-附件-01_AI未解析 文件夹为空, 无需更新")
            else:
                for f in os.listdir(os.path.join(keyword, '附件', '01_AI未解析')):
                    item_path = os.path.join(keyword, '附件', '01_AI未解析', f)
                    if os.path.isdir(item_path):
                        logger.info(f"{f} 是文件夹,开始处理")
                        try:
                            # AI理解docx内容
                            for pdf_file in os.listdir(item_path):
                                if (pdf_file.endswith('.pdf') or pdf_file.endswith('.docx')) and '商务' in pdf_file.split('.')[0]:
                                    md = MarkItDown(docintel_endpoint="<document_intelligence_endpoint>")
                                    result = md.convert(os.path.join(item_path, pdf_file))
                                    input_text = result.text_content[:19999]
                                    shangwu_result_format = f"{separator}\n# 《{pdf_file}》\n{input_text}\n\n\n"
                                elif (pdf_file.endswith('.pdf') or pdf_file.endswith('.docx')) and '技术' in pdf_file.split('.')[0]:
                                    md = MarkItDown(docintel_endpoint="<document_intelligence_endpoint>")
                                    result = md.convert(os.path.join(item_path, pdf_file))
                                    input_text = result.text_content[:19999]
                                    jishu_result_format = f"{separator}\n# 《{pdf_file}》\n{input_text}"
                            result_format = shangwu_result_format + jishu_result_format
                            """
                            步骤2：调用工作流API，传入文件ID和输入文本
                            """
                            url = f"{base_url}/api/v1/chat/completions"
                            headers = {
                                'Authorization': f'Bearer {api_key}',
                                'Content-Type': 'application/json'
                            }
                            # 构建请求数据体
                            data = {
                                "model": "fastgpt-workflow",  # 或者其他指定的模型名
                                "chatId": chat_id,  # 用于保持会话的连续性:cite[9]
                                "workflowId": workflow_id,  # 指定要运行的工作流
                                "messages": [
                                    {
                                        "role": "user",
                                        "content": result_format,
                                        # 此处是关键：在消息中关联已上传的文件
                                        "files": [file_id]  # 假设API支持通过`files`字段传递文件ID列表
                                    }
                                ]
                            }
                            try:
                                response = requests.post(url, json=data, headers=headers)
                                response.raise_for_status()
                                result = response.json()
                                logger.info("工作流调用成功！")
                                matches = re.findall(pattern, result["choices"][0]["message"]["content"])
                                json_str = matches[0].replace('\n', '')
                                json_dict = json.loads(json_str)
                                # 写入到excel文件
                                write_excel2((f), json_dict, keyword)
                            except Exception as e:
                                logger.info(f"工作流调用失败: {e}")
                                logger.info(f"响应状态码: {response.status_code}")
                                logger.info(f"响应内容: {response.text}")

                            logger.info(f'{keyword} - 数据更新成功')

                            if '不可承接' in json_dict['业务承接判定']:
                                if not os.path.exists(os.path.join(keyword, '附件', '02_AI已解析(不可承接)')):
                                    os.makedirs(os.path.join(keyword, '附件', '02_AI已解析(不可承接)'))
                                shutil.move(os.path.join(keyword, '附件', '01_AI未解析', f),
                                            os.path.join(keyword, '附件', '02_AI已解析(不可承接)'))
                            else:
                                if not os.path.exists(os.path.join(keyword, '附件', '03_AI已解析(可承接)')):
                                    os.makedirs(os.path.join(keyword, '附件', '03_AI已解析(可承接)'))
                                shutil.move(os.path.join(keyword, '附件', '01_AI未解析', f),
                                            os.path.join(keyword, '附件', '03_AI已解析(可承接)'))

                        except Exception as e:
                            logger.error(f"AI文档解析文件夹内容时发生错误: {e}")
                            continue

                    elif os.path.isfile(item_path):
                        if f.lower().endswith('.zip'):
                            logger.info(f"{f} 是ZIP文件, 开始处理")
                            try:
                                # ai阅读理解pdf文件
                                with zipfile.ZipFile(os.path.join(keyword, '附件', '01_AI未解析', f), 'r') as zip_ref:
                                    zip_ref.extractall("临时文件")
                                    logger.info(f"成功解压到: {"临时文件"}")
                                # AI理解pdf、docx内容
                                for pdf_file in os.listdir('临时文件'):
                                    if pdf_file.endswith('.pdf') and '商务' in pdf_file.split('.')[0]:
                                        md = MarkItDown(docintel_endpoint="<document_intelligence_endpoint>")
                                        result = md.convert('临时文件' + '/' + pdf_file)
                                        input_text = result.text_content[:19999]
                                        shangwu_result_format = f"{separator}\n# 《{pdf_file}》\n{input_text}\n\n\n"
                                    elif pdf_file.endswith('.pdf') and '技术' in pdf_file.split('.')[0]:
                                        md = MarkItDown(docintel_endpoint="<document_intelligence_endpoint>")
                                        result = md.convert('临时文件' + '/' + pdf_file)
                                        input_text = result.text_content[:19999]
                                        jishu_result_format = f"{separator}\n# 《{pdf_file}》\n{input_text}"
                                result_format = shangwu_result_format + jishu_result_format
                                """
                                步骤2：调用工作流API，传入文件ID和输入文本
                                """
                                url = f"{base_url}/api/v1/chat/completions"
                                headers = {
                                    'Authorization': f'Bearer {api_key}',
                                    'Content-Type': 'application/json'
                                }
                                # 构建请求数据体
                                data = {
                                    "model": "fastgpt-workflow",  # 或者其他指定的模型名
                                    "chatId": chat_id,  # 用于保持会话的连续性:cite[9]
                                    "workflowId": workflow_id,  # 指定要运行的工作流
                                    "messages": [
                                        {
                                            "role": "user",
                                            "content": result_format,
                                            # 此处是关键：在消息中关联已上传的文件
                                            "files": [file_id]  # 假设API支持通过`files`字段传递文件ID列表
                                        }
                                    ]
                                }

                                try:
                                    response = requests.post(url, json=data, headers=headers)
                                    response.raise_for_status()
                                    result = response.json()
                                    logger.info("工作流调用成功！")
                                    matches = re.findall(pattern, result["choices"][0]["message"]["content"])
                                    json_str = matches[0].replace('\n', '')
                                    json_dict = json.loads(json_str)
                                    # 写入到excel文件
                                    write_excel2((f), json_dict, keyword)
                                except Exception as e:
                                    logger.info(f"工作流调用失败: {e}")
                                    logger.info(f"响应状态码: {response.status_code}")
                                    logger.info(f"响应内容: {response.text}")

                                # 4. 删除临时文件中的所有文件
                                for file in os.listdir('临时文件'):
                                    os.remove('临时文件' + '/' + file)

                                logger.info(f'{keyword} - 数据更新成功')

                                if '不可承接' in json_dict['业务承接判定']:
                                    if not os.path.exists(os.path.join(keyword, '附件', '02_AI已解析(不可承接)')):
                                        os.makedirs(os.path.join(keyword, '附件', '02_AI已解析(不可承接)'))
                                    shutil.move(os.path.join(keyword, '附件', '01_AI未解析', f),
                                                os.path.join(keyword, '附件', '02_AI已解析(不可承接)'))
                                else:
                                    if not os.path.exists(os.path.join(keyword, '附件', '03_AI已解析(可承接)')):
                                        os.makedirs(os.path.join(keyword, '附件', '03_AI已解析(可承接)'))
                                    shutil.move(os.path.join(keyword, '附件', '01_AI未解析', f),
                                                os.path.join(keyword, '附件', '03_AI已解析(可承接)'))

                            except Exception as e:
                                logger.error(f"AI文档解析zip内容时发生错误: {e}")
                                continue
                        else:
                            logger.info(f"{f} 是其他类型文件,暂不处理")
                format_excel_file(os.path.join(os.getcwd(), keyword) + '/' + '02_招标文件智能解析结果.xlsx')




if __name__ == '__main__':

    keywords_list = ['软件', '运维', '维保']
    try:
        main(keywords_list)
        main2(keywords_list)
    except KeyboardInterrupt:
        logger.info("程序被用户中断")
    except Exception as e:
        logger.error(f"程序异常退出: {str(e)}")
        sys.exit(1)
